import csv
import inspect
import logging
import softest
from openpyxl import load_workbook

class Utils(softest.TestCase):

    def assert_text_of_web_elem_list(self,lst,expected_text):
        print(expected_text)
        for item in lst:
            print(item.text)
            self.soft_assert(self.assertEqual,item.text,expected_text)
            if item.text == expected_text:
                print("Test Case Passed")
            else:
                print("Test Case Failed")
        self.assert_all()

    def cust_logger(loglevel=logging.DEBUG):

        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(loglevel)
        formatter = logging.Formatter("%(asctime)s - %(levelname)s- %(name)s - %(message)s",datefmt="%m-%d-%Y %H:%M:%S")
        fh = logging.FileHandler("automation.log","a")
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger

    def excel_reader(file_path,sheet_name):

        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        max_rows = ws.max_row
        max_columns = ws.max_column

        outer_list = []
        inner_list = []
        for i in range(2,max_rows+1):
            for j in range(1,max_columns+1):
                inner_list.append(ws.cell(i,j).value)
            outer_list.append(inner_list.copy())
            inner_list.clear()
        return outer_list

    def csv_reader(file_path):
        data_list = []
        with open(file_path, newline='',encoding='utf-8') as f:
            reader = csv.reader(f)
            i = 0
            for row in reader:
                if i == 0:
                    i += 1
                else:
                    data_list.append(row)
        return  data_list
